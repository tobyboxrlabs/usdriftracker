#!/usr/bin/env python3
"""
Excel export script for USDRIF Mint/Redeem Analyser
Uses openpyxl for formatting with Font, Alignment, and PatternFill
"""

import json
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def export_to_excel(transactions_data, output_path):
    """
    Export transactions to Excel with formatting
    
    Args:
        transactions_data: List of transaction dictionaries
        output_path: Path to save the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "USDRIF Mint Redeem Transactions"  # Sheet names cannot contain : \ / ? * [ ]
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Column headers
    headers = [
        "Time (UTC)",
        "Status",
        "Asset",
        "Type",
        "Amount Minted/Redeemed",
        "Receiver",
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data rows
    for row_idx, tx in enumerate(transactions_data, start=2):
        # Time (UTC)
        ws.cell(row=row_idx, column=1).value = tx.get('time', '')
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
        
        # Status
        status_cell = ws.cell(row=row_idx, column=2)
        status_cell.value = tx.get('status', '')
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        if tx.get('status') == 'Success':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif tx.get('status') == 'Failed':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Asset
        ws.cell(row=row_idx, column=3).value = tx.get('asset', '')
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        
        # Type
        type_cell = ws.cell(row=row_idx, column=4)
        type_cell.value = tx.get('type', '')
        type_cell.alignment = Alignment(horizontal="center", vertical="center")
        if 'Mint' in tx.get('type', ''):
            type_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        elif 'Redeem' in tx.get('type', ''):
            type_cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        # Amount Minted/Redeemed
        amount_cell = ws.cell(row=row_idx, column=5)
        amount_cell.value = tx.get('amountMintedRedeemed', '')
        amount_cell.alignment = Alignment(horizontal="right", vertical="center")
        amount_cell.number_format = '#,##0'
        
        # Receiver
        receiver_cell = ws.cell(row=row_idx, column=6)
        receiver_cell.value = tx.get('receiver', '')
        receiver_cell.alignment = Alignment(horizontal="left", vertical="center")
        receiver_cell.font = Font(name="Courier New", size=9)  # Monospace for addresses
    
    # Auto-adjust column widths
    column_widths = {
        'A': 20,  # Time
        'B': 12,  # Status
        'C': 12,  # Asset
        'D': 15,  # Type
        'E': 25,  # Amount
        'F': 45,  # Receiver
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add summary row at the bottom
    summary_row = len(transactions_data) + 3
    ws.cell(row=summary_row, column=1).value = "Total Transactions:"
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row, column=2).value = len(transactions_data)
    ws.cell(row=summary_row, column=2).font = Font(bold=True)
    
    mints = sum(1 for tx in transactions_data if 'Mint' in tx.get('type', ''))
    redeems = sum(1 for tx in transactions_data if 'Redeem' in tx.get('type', ''))
    
    ws.cell(row=summary_row + 1, column=1).value = "Mints:"
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2).value = mints
    
    ws.cell(row=summary_row + 2, column=1).value = "Redeems:"
    ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2).value = redeems
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file saved to: {output_path}")

def main():
    """Main function to handle command-line usage"""
    if len(sys.argv) < 3:
        print("Usage: python export_to_excel.py <input_json_file> <output_xlsx_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2]
    
    # Read JSON data
    with open(input_file, 'r') as f:
        transactions_data = json.load(f)
    
    # Export to Excel
    export_to_excel(transactions_data, output_file)

if __name__ == "__main__":
    main()
